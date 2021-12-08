# make a list of all roots + sign + meaning
# make a list of all root families

# make a list of all the copound families
# sort the list alphebetically

# alternative method:
# work directly on the xlxs
# remove all rows without family
# sort by root > number > meaning > family > headword
# delete unused columns
# save as csv

# ignore help etc "does not contain √"

import csv
# from os import close, name
import openpyxl as xl
import re
from xlsxwriter.workbook import Workbook

# csv_to_excel
print("converting csv to excel")

csv_file = "/home/bhikkhu/Bodhirasa/Dropbox/Pāli English Dictionary/Pāli English Dictionary-full.csv"
xlsx_file = "Pāli English Dictionary-full.xlsx"

workbook = Workbook(xlsx_file)
worksheet = workbook.add_worksheet()

csv_reader = csv.reader(open(csv_file,'rt'),delimiter="\t")

for row, data in enumerate(csv_reader):
    worksheet.write_row(row, 0, data)

workbook.close()

# open excel sheet
print("loading excel sheet")

wb = xl.load_workbook("Pāli English Dictionary-full.xlsx")
sheet = wb["Sheet1"]
last_row = sheet.max_row + 1
row_number = 2
root_famiy_list = []

# extract root families
print("extracting root family names")

for row in range(row_number, last_row):
    family_cell = sheet.cell(row, 25)
    if family_cell.value != None:
        root_famiy_list.append(family_cell.value)

root_famiy_list.sort()
root_famiy_list = list(dict.fromkeys(root_famiy_list))

# write families
print("writing families")

line_break = ("~" * 40)
print(line_break)

txt_file = open ("root_families.csv", 'w', encoding= "'utf-8")
txt_file.write (f"root families\n")

family_count = 0

for family in root_famiy_list:
    txt_file.write (f"\n")
    txt_file.write (f"{family}\n")
    
    for row in range(2, last_row):
        family_cell = sheet.cell(row, 25)
        headword = sheet.cell(row, 1)
        pos = sheet.cell(row, 4)
        meaning = sheet.cell(row, 11)
        buddhadatta = sheet.cell(row, 50)
        if family_cell.value == family:
            txt_file.write (f"{headword.value}\t{pos.value}\t{meaning.value}\t{buddhadatta.value}\n")
            family_count += 1
            print(f"{family_count}. {family}")
        else:
            continue

print(f"{family_count} familes")
print(f"saved to {txt_file.name}")
print("fin")


# 1	Pāli1
# 2	Pāli2
# 3	Fin
# 4	POS
# 5	Grammar
# 6	Derived from
# 7	Neg
# 8	Verb
# 9	Trans
# 10	Case
# 11	Meaning IN CONTEXT
# 12	Literal Meaning
# 13	Non IA
# 14	Sanskrit
# 15	Sk Root
# 16	Sk Root Mn
# 17	Cl
# 18	Pāli Root
# 19	Root In Comps
# 20	V
# 21	Grp
# 22	Sgn
# 23	Root Meaning
# 24	Base
# 25	Family
# 26	Family2
# 27	Construction
# 28	Derivative
# 29	Suffix
# 30	Phonetic Changes
# 31	Compound
# 32	Compound Construction
# 33	Non-Root In Comps
# 34	Source1
# 35	Sutta1
# 36	Example1
# 37	Source 2
# 38	Sutta2
# 39	Example 2
# 40	Antonyms
# 41	Synonyms – different word
# 42	Variant – same constr or diff reading
# 43	Commentary
# 44	Notes
# 45	Cognate
# 46	Category
# 47	Link
# 48	Stem
# 49	Pattern
# 50	Buddhadatta
# 51	22
# 52	Pāli1 ≠ const
# 53	test dupl
# 54	Metadata